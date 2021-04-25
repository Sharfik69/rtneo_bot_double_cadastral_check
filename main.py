from openpyxl import load_workbook, Workbook


class Checker:

    def __init__(self, file_name, sheet_name, kn, street_, house_, flat_):

        self.wb_name = file_name.replace('.xlsx', '')
        self.sheet_name = sheet_name
        self.street = street_
        self.house = house_
        self.flat = flat_
        self.kn = kn
        self.answer = '{} проверка.xlsx'.format(self.wb_name)

        self.wb1 = Workbook()
        self.ss = self.wb1.active

        self.a = set()
        self.d = {}

    def load_book(self):
        try:
            self.wb = load_workbook('{}.xlsx'.format(self.wb_name))
            self.s = self.wb[self.sheet_name]
            return True
        except Exception:
            return False

    def check(self):
        for i in range(2, 400000):
            if self.s.cell(row=i, column=1).value is None:
                break
            kn = self.s.cell(row=i, column=self.kn + 1).value
            if kn is not None:
                kn = str(kn)
            if kn is not None and kn[0].isdigit():
                info = []

                for j in range(1, 40):
                    info.append(self.s.cell(row=i, column=j).value)

                if kn not in self.d:
                    self.d[kn] = []

                self.d[kn].append(info)

        row_ = 1
        ans = 0

        for kn in self.d:
            if len(self.d[kn]) == 1:
                continue

            street_set, house_set, apart_set = set(), set(), set()

            for j in self.d[kn]:
                street_set.add(j[self.street].lower().replace('-й', ''))
                house_set.add(j[self.house])
                apart_set.add(j[self.flat])

            if len(street_set) == 1 and len(house_set) == 1 and len(apart_set) == 1:
                continue

            ans += 1
            for j in range(len(self.d[kn])):
                for col_ in range(len(self.d[kn][j])):
                    self.ss.cell(row=row_, column=col_ + 1).value = self.d[kn][j][col_]
                row_ += 1
            row_ += 1

        ss = self.wb1.create_sheet("список кадастров")

        row_ = 1
        for i in self.d.keys():
            if len(self.d[i]) == 1:
                continue
            ss.cell(row=row_, column=1).value = i
            row_ += 1
        self.wb1.save(self.answer)
        self.cnt = ans
        return True

    def get_file_name(self):
        return self.answer